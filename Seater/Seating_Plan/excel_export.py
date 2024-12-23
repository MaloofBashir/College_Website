import PyPDF2
import openpyxl
import sys
from django.http import HttpResponse

args=sys.argv


mystr=""
remove_subs=list(mystr.split('\t'))
#print((remove_subs))

all_sub_names=[]

#Adding Subject from all pages to the one list
def add_to_allsubs(subjects):
      for sub in subjects:
            if sub not in all_sub_names:
                  if sub not in remove_subs:
                        all_sub_names.append(sub)



all_roll_nos=[]

#Getting Roll NO of the Particular Page NO/Student
def get_roll_no(pdf_obj,page_no):
      page1=pdf_obj.pages[page_no].extract_text()
      if page1:
            lines=page1.split('\n')

            for lineno, line in enumerate(lines):
                  if lineno==14:
                        rollno=line.split(".")[1]
                        return rollno

#Getting all Subject Names of individual Student/Page
def get_subname(pdf_obj,page_no):
        page1=pdf_obj.pages[page_no].extract_text() 

        if page1:
            lines=page1.split('\n')

            for lineno,line in enumerate(lines):
                        
                if lineno==19:
                        subjects=line.split(" ")[0]
                        subjects=subjects.split("-")
                        add_to_allsubs(subjects)
                        return subjects

wb=openpyxl.Workbook()
ws=wb.active





#Adding Roll NOs of a subject to active sheet
def add_subject_to_sheet(subject,col_number,pdffile):
    
        pdf_obj=PyPDF2.PdfReader(pdffile)
        num_pages=len(pdf_obj.pages)
        sub_to_search=subject
        rowno=1
        ws.cell(row=rowno,column=col_number,value=subject)
        

        for pageno in range(num_pages):
                subs_of_current_candidate=get_subname(pdf_obj,pageno)

                if sub_to_search in subs_of_current_candidate:

                    rowno=rowno+1
                    cur_roll_no=get_roll_no(pdf_obj,pageno)
                    ws.cell(row=rowno,column=col_number,value=cur_roll_no) 

                    all_roll_nos.append(cur_roll_no)  



def getting_all_subs(pdffile):
        pdf_obj=PyPDF2.PdfReader(pdffile)
        num_pages=len(pdf_obj.pages)
        for pageno in range(num_pages):
            subs_of_current_candidate=get_subname(pdf_obj,pageno)

#looping through all subject in the document one by one
#This if the first function to run


# def excel_export(file):
#     print(file)
#     getting_all_subs(file)
#     for colno,subject in enumerate(all_sub_names):
#         add_subject_to_sheet(subject,colno+1,file)
#     excel_filename="myfile.xlsx"
#     wb.save(excel_filename)
#     return excel_filename

import openpyxl
import io

def excel_export(file):
    # Process the PDF file (you are already extracting subjects and roll numbers)
    getting_all_subs(file)

    # Create a new Excel workbook
    # Add the data to the Excel file (e.g., subjects and roll numbers)
    for colno, subject in enumerate(all_sub_names):
        add_subject_to_sheet(subject, colno + 1, file)

    # Save the workbook to an in-memory buffer (instead of saving to disk)
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)  # Rewind the buffer to the beginning

    # Prepare the HTTP response to download the Excel file
    response = HttpResponse(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=myfile.xlsx'  # Set the filename for download

    return response
